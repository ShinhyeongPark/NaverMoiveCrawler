# To show progress bar
from tqdm import tqdm
import check
import excel as xl
import google_nl as g
import print_results
import RepleCrawler

from PyQt5.QtWidgets import QMainWindow, QApplication, QLabel, QVBoxLayout, QPushButton
import pandas as pd
import matplotlib.pyplot as plt
from PyQt5.QtWidgets import QDialog

import requests
import openpyxl
import operator
from openpyxl import Workbook
from bs4 import BeautifulSoup as bs



def createExcelReple(title):
    # titleNurl = SearchingPage.InputTitle()  # str형태 타이틀과 url 반환

    wb = openpyxl.load_workbook('2019년 영화.xlsx')
    ws = wb.active
    for i in ws.rows:
        if (operator.eq(i[0].value, title)):
            titleNurl = str(title+i[1].value)

    endpoint = titleNurl.find("https")
    title=titleNurl[0:endpoint]
    url=titleNurl[endpoint:len(titleNurl)]
    print(title)
    print(url)

    html = bs(requests.get(url).content, "html.parser",from_encoding='utf-8')
    cnt=html.select("body > div > div > div.score_total > strong > em")[0].contents[0].replace(',','')

    wb=Workbook()
    ws=wb.create_sheet(title,0)
    row=2
    ws.cell(1,1,"평점")
    ws.cell(1,2,"좋아요")
    ws.cell(1,3,"싫어요")
    ws.cell(1,4,"비율")
    ws.cell(1,5,"댓글")


    for x in range(1, int(cnt)//10+1):
         html=bs(requests.get(url+"&page="+str(x)).content, "html.parser",from_encoding="utf-8")
         score = html.select("body > div > div > div.score_result > ul > li > div.star_score > em")
         reple = html.select("body > div > div > div.score_result > ul > li > div.score_reple > p")
         like = html.select("body > div > div> div.score_result > ul > li > div.btn_area")
         for i in range(len(reple)):

              ws.cell(row, 5,int(score[i].contents[0])) #평점
              ws.cell(row, 2, int(like[i].contents[1].contents[5].contents[0])) #좋아요
              ws.cell(row, 3, int(like[i].contents[3].contents[5].contents[0])) #싫어요
              if int(like[i].contents[1].contents[5].contents[0])==0 or \
                    int(like[i].contents[3].contents[5].contents[0])==0:
                    ws.cell(row,4,1)
              else:
                    ws.cell(row,4,int(like[i].contents[1].contents[5].contents[0]) /
                    int(like[i].contents[3].contents[5].contents[0]))
              tmp = reple[i].contents[5].contents[0].strip()
              if tmp == "":
                    try:
                          tmp=reple[i].contents[5].contents[1].contents[1]['data-src']
                    except:
                          continue
              ws.cell(row, 1, tmp)
              row += 1
    wb.save('C:/Users/홍석찬/Desktop/nlp-master/' + title + '.xlsx')

    FILE =('C:/Users/홍석찬/Desktop/nlp-master/' + title + '.xlsx')
    NEW_FILE =('C:/Users/홍석찬/Desktop/nlp-master/' + title + '_.xlsx')
    TYPE =("excel")
   # FILE, NEW_FILE, TYPE =('C:/Users/홍석찬/Desktop/nlp-master/' + title + '.xlsx')

    DATA, ROWS_NB = xl.get_xl(FILE, NEW_FILE)
    print(range(ROWS_NB))
    print(tqdm(range(ROWS_NB)));
    for ROW in tqdm(range(ROWS_NB)):
        print(ROW)
        print("######")
        print(DATA.iloc[ROW, 0])
        print(DATA.iloc[ROW, 1])

        SEN = g.analyze_sentiment(DATA.iloc[ROW, 0])
        print(SEN)
        DATA.iloc[ROW, 1] = round(SEN.score, 1) * 10
        DATA.iloc[ROW, 2] = round(SEN.magnitude, 1)
    DATA.to_excel(NEW_FILE, index=None)
    print_results.excel(NEW_FILE)

    df = pd.read_excel('C:/Users/홍석찬/Desktop/nlp-master/' + title + '_.xlsx')
    dx = df[['Score']]

    class MyDialog(QDialog):

        plt.plot(dx)
        plt.show()


    app = QApplication([])
    dialog = MyDialog()
    dialog.show()
    app.exec_()





