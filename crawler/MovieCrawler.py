import requests
import re
from openpyxl import Workbook
from bs4 import BeautifulSoup as bs

url = "https://movie.naver.com/movie/sdb/browsing/bmovie.nhn?open=2019"
html = bs(requests.get(url).content, "html.parser",from_encoding='utf-8')
cnt = 882


wb=Workbook()
ws=wb.create_sheet("2019년 영화",0)
row=2
ws.cell(1,1,"영화제목")
ws.cell(1,2,"URL")

for x in range(1, int(cnt)//20+1):
      html=bs(requests.get(url+"&page="+str(x)).content, "html.parser",from_encoding="utf-8")
      for i in range(0,19):
            title = html.select("#old_content > ul > li >a")[i]
            tt=str(title) #title을 str로 변환 추출을 위해
            code=re.findall("\d+",tt[37:43])[0] #영화 코드번호 추출
            endpoint = title.contents[0].find(" (") #제목의 끝점
            if(endpoint >= 0):
                real_title=title.contents[0][0:endpoint]
            else:
                real_title=title.contents[0]

            ws.cell(row, 1, real_title)
            ws.cell(row, 2, "https://movie.naver.com/movie/bi/mi/pointWriteFormList.nhn?code="+code+
                    "&type=after&onlyActualPointYn=Y&onlySpoilerPointYn=N&order=sympathyScore")

            row += 1

wb.save('2019년 영화.xlsx')